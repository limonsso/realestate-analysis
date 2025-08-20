#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
💾 EXPORTEUR AVANCÉ - MULTI-FORMATS ET OPTIMISATIONS
=====================================================

Module d'export avancé avec support multi-formats
Basé sur les spécifications du real_estate_prompt.md
"""

import pandas as pd
import numpy as np
import logging
from typing import Dict, List, Tuple, Optional, Any, Union
from pathlib import Path
import json
import pickle
import warnings
from datetime import datetime
import os

# Imports conditionnels pour les formats spéciaux
try:
    import geopandas as gpd
    from shapely.geometry import Point
    GEOPANDAS_AVAILABLE = True
except ImportError:
    GEOPANDAS_AVAILABLE = False
    warnings.warn("GeoPandas non disponible - export GeoJSON limité")

try:
    import h5py
    H5PY_AVAILABLE = True
except ImportError:
    H5PY_AVAILABLE = False
    warnings.warn("H5Py non disponible - export HDF5 limité")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn("OpenPyXL non disponible - export Excel limité")

warnings.filterwarnings('ignore')
logger = logging.getLogger(__name__)

class AdvancedExporter:
    """
    Exporteur avancé avec support multi-formats et optimisations
    Parquet, CSV, GeoJSON, HDF5, Excel, JSON, Pickle
    """
    
    def __init__(self, export_config: Dict = None):
        """
        Initialise l'exporteur avancé
        
        Args:
            export_config: Configuration d'export
        """
        self.export_config = export_config or self._default_export_config()
        self.export_history = []
        self.export_stats = {}
        
        logger.info("💾 AdvancedExporter initialisé")
        logger.info(f"🌍 GeoPandas: {'✅' if GEOPANDAS_AVAILABLE else '❌'}")
        logger.info(f"📊 H5Py: {'✅' if H5PY_AVAILABLE else '❌'}")
        logger.info(f"📈 OpenPyXL: {'✅' if OPENPYXL_AVAILABLE else '❌'}")
    
    def _default_export_config(self) -> Dict:
        """Configuration d'export par défaut"""
        return {
            "formats": ["parquet", "csv", "geojson", "hdf5"],
            "compression": "snappy",  # Compression Parquet
            "encoding": "utf-8",
            "float_format": "%.6f",
            "index": False,
            "output_directory": "exports",
            "filename_prefix": "real_estate_data",
            "timestamp_format": "%Y%m%d_%H%M%S",
            "chunk_size": 10000,  # Pour les gros datasets
            "parallel_export": True,
            "memory_optimization": True
        }
    
    def export_dataset(self, df: pd.DataFrame, dataset_name: str = "dataset", 
                      formats: List[str] = None, output_dir: str = None) -> Dict[str, str]:
        """
        Exporte le dataset dans plusieurs formats
        
        Args:
            df: DataFrame à exporter
            dataset_name: Nom du dataset
            formats: Formats d'export (si None, utilise la config par défaut)
            output_dir: Répertoire de sortie (si None, utilise la config par défaut)
            
        Returns:
            Dict avec les chemins des fichiers exportés
        """
        logger.info(f"💾 === EXPORT MULTI-FORMATS: {dataset_name} ===")
        
        # Configuration
        export_formats = formats or self.export_config["formats"]
        output_directory = output_dir or self.export_config["output_directory"]
        
        # Création du répertoire de sortie
        Path(output_directory).mkdir(parents=True, exist_ok=True)
        
        # Génération du timestamp
        timestamp = datetime.now().strftime(self.export_config["timestamp_format"])
        
        # Export dans chaque format
        exported_files = {}
        export_start = datetime.now()
        
        for format_type in export_formats:
            try:
                logger.info(f"📤 Export {format_type.upper()}...")
                
                if format_type == "parquet":
                    file_path = self._export_parquet(df, dataset_name, timestamp, output_directory)
                elif format_type == "csv":
                    file_path = self._export_csv(df, dataset_name, timestamp, output_directory)
                elif format_type == "geojson":
                    file_path = self._export_geojson(df, dataset_name, timestamp, output_directory)
                elif format_type == "hdf5":
                    file_path = self._export_hdf5(df, dataset_name, timestamp, output_directory)
                elif format_type == "excel":
                    file_path = self._export_excel(df, dataset_name, timestamp, output_directory)
                elif format_type == "json":
                    file_path = self._export_json(df, dataset_name, timestamp, output_directory)
                elif format_type == "pickle":
                    file_path = self._export_pickle(df, dataset_name, timestamp, output_directory)
                else:
                    logger.warning(f"⚠️ Format non supporté: {format_type}")
                    continue
                
                if file_path:
                    exported_files[format_type] = file_path
                    logger.info(f"✅ {format_type.upper()} exporté: {file_path}")
                
            except Exception as e:
                logger.error(f"❌ Erreur export {format_type}: {e}")
                exported_files[format_type] = f"ERROR: {str(e)}"
        
        # Calcul des statistiques
        export_end = datetime.now()
        export_duration = export_end - export_start
        
        # Mise à jour de l'historique
        export_record = {
            "timestamp": export_start.isoformat(),
            "dataset_name": dataset_name,
            "formats": export_formats,
            "exported_files": exported_files,
            "duration_seconds": export_duration.total_seconds(),
            "dataset_shape": df.shape,
            "memory_usage_mb": df.memory_usage(deep=True).sum() / 1024 / 1024
        }
        
        self.export_history.append(export_record)
        
        # Statistiques d'export
        successful_exports = len([f for f in exported_files.values() if not str(f).startswith("ERROR")])
        self.export_stats[dataset_name] = {
            "total_formats": len(export_formats),
            "successful_exports": successful_exports,
            "success_rate": successful_exports / len(export_formats) if export_formats else 0,
            "last_export": export_start.isoformat()
        }
        
        logger.info(f"🎯 Export terminé en {export_duration.total_seconds():.2f}s")
        logger.info(f"📊 {successful_exports}/{len(export_formats)} formats exportés avec succès")
        
        return exported_files
    
    def _export_parquet(self, df: pd.DataFrame, dataset_name: str, 
                        timestamp: str, output_dir: str) -> str:
        """Export au format Parquet optimisé"""
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.parquet"
            file_path = os.path.join(output_dir, filename)
            
            # Optimisation des types de données pour Parquet
            optimized_df = self._optimize_dataframe_for_parquet(df)
            
            # Export avec compression
            optimized_df.to_parquet(
                file_path,
                compression=self.export_config["compression"],
                index=self.export_config["index"],
                engine='pyarrow'
            )
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export Parquet: {e}")
            return None
    
    def _export_csv(self, df: pd.DataFrame, dataset_name: str, 
                    timestamp: str, output_dir: str) -> str:
        """Export au format CSV optimisé"""
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.csv"
            file_path = os.path.join(output_dir, filename)
            
            # Export avec optimisations
            df.to_csv(
                file_path,
                encoding=self.export_config["encoding"],
                float_format=self.export_config["float_format"],
                index=self.export_config["index"]
            )
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export CSV: {e}")
            return None
    
    def _export_geojson(self, df: pd.DataFrame, dataset_name: str, 
                        timestamp: str, output_dir: str) -> str:
        """Export au format GeoJSON pour données géospatiales"""
        if not GEOPANDAS_AVAILABLE:
            logger.warning("⚠️ GeoPandas non disponible - export GeoJSON ignoré")
            return None
        
        try:
            # Détection des colonnes géographiques
            lat_cols = [col for col in df.columns if any(term in col.lower() for term in ['lat', 'latitude'])]
            lng_cols = [col for col in df.columns if any(term in col.lower() for term in ['lng', 'long', 'longitude'])]
            
            if not lat_cols or not lng_cols:
                logger.warning("⚠️ Aucune colonne géographique détectée - export GeoJSON ignoré")
                return None
            
            # Utilisation des premières colonnes géographiques trouvées
            lat_col = lat_cols[0]
            lng_col = lng_cols[0]
            
            # Création de la géométrie
            geometry = [Point(xy) for xy in zip(df[lng_col], df[lat_col])]
            
            # Création du GeoDataFrame
            gdf = gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")
            
            # Export
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.geojson"
            file_path = os.path.join(output_dir, filename)
            
            gdf.to_file(file_path, driver='GeoJSON')
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export GeoJSON: {e}")
            return None
    
    def _export_hdf5(self, df: pd.DataFrame, dataset_name: str, 
                     timestamp: str, output_dir: str) -> str:
        """Export au format HDF5 optimisé"""
        if not H5PY_AVAILABLE:
            logger.warning("⚠️ H5Py non disponible - export HDF5 ignoré")
            return None
        
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.h5"
            file_path = os.path.join(output_dir, filename)
            
            # Export avec optimisations
            df.to_hdf(
                file_path,
                key='data',
                mode='w',
                format='table',
                complevel=9,  # Compression maximale
                complib='blosc'
            )
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export HDF5: {e}")
            return None
    
    def _export_excel(self, df: pd.DataFrame, dataset_name: str, 
                      timestamp: str, output_dir: str) -> str:
        """Export au format Excel avec formatage"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("⚠️ OpenPyXL non disponible - export Excel ignoré")
            return None
        
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.xlsx"
            file_path = os.path.join(output_dir, filename)
            
            # Export avec optimisations
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(
                    writer,
                    sheet_name='Data',
                    index=self.export_config["index"],
                    float_format=self.export_config["float_format"]
                )
                
                # Formatage de la feuille
                worksheet = writer.sheets['Data']
                self._format_excel_worksheet(worksheet, df)
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export Excel: {e}")
            return None
    
    def _export_json(self, df: pd.DataFrame, dataset_name: str, 
                     timestamp: str, output_dir: str) -> str:
        """Export au format JSON optimisé"""
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.json"
            file_path = os.path.join(output_dir, filename)
            
            # Conversion en format JSON optimisé
            json_data = {
                "metadata": {
                    "export_timestamp": timestamp,
                    "dataset_name": dataset_name,
                    "shape": df.shape,
                    "columns": list(df.columns)
                },
                "data": df.to_dict('records')
            }
            
            # Export avec indentation
            with open(file_path, 'w', encoding=self.export_config["encoding"]) as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export JSON: {e}")
            return None
    
    def _export_pickle(self, df: pd.DataFrame, dataset_name: str, 
                       timestamp: str, output_dir: str) -> str:
        """Export au format Pickle pour Python"""
        try:
            filename = f"{self.export_config['filename_prefix']}_{dataset_name}_{timestamp}.pkl"
            file_path = os.path.join(output_dir, filename)
            
            # Export avec compression
            df.to_pickle(file_path, compression='gzip')
            
            return file_path
            
        except Exception as e:
            logger.error(f"❌ Erreur export Pickle: {e}")
            return None
    
    def _optimize_dataframe_for_parquet(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimise le DataFrame pour l'export Parquet"""
        optimized_df = df.copy()
        
        # Optimisation des types numériques
        for col in optimized_df.select_dtypes(include=[np.number]).columns:
            if optimized_df[col].dtype == 'float64':
                # Conversion en float32 si possible
                if optimized_df[col].notna().all():
                    min_val = optimized_df[col].min()
                    max_val = optimized_df[col].max()
                    if min_val >= -3.4e38 and max_val <= 3.4e38:
                        optimized_df[col] = optimized_df[col].astype('float32')
            
            elif optimized_df[col].dtype == 'int64':
                # Conversion en int32 si possible
                if optimized_df[col].notna().all():
                    min_val = optimized_df[col].min()
                    max_val = optimized_df[col].max()
                    if min_val >= -2147483648 and max_val <= 2147483647:
                        optimized_df[col] = optimized_df[col].astype('int32')
        
        # Optimisation des types catégoriels
        for col in optimized_df.select_dtypes(include=['object']).columns:
            if optimized_df[col].nunique() / len(optimized_df) < 0.5:  # Moins de 50% de valeurs uniques
                optimized_df[col] = optimized_df[col].astype('category')
        
        return optimized_df
    
    def _format_excel_worksheet(self, worksheet, df: pd.DataFrame):
        """Formate la feuille Excel pour une meilleure lisibilité"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Application des styles aux en-têtes
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Ajustement automatique de la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Limite à 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"⚠️ Formatage Excel limité: {e}")
    
    def export_with_metadata(self, df: pd.DataFrame, dataset_name: str, 
                            metadata: Dict = None, output_dir: str = None) -> Dict[str, str]:
        """
        Export avec métadonnées enrichies
        
        Args:
            df: DataFrame à exporter
            dataset_name: Nom du dataset
            metadata: Métadonnées supplémentaires
            output_dir: Répertoire de sortie
            
        Returns:
            Dict avec les chemins des fichiers exportés
        """
        logger.info(f"📊 === EXPORT AVEC MÉTADONNÉES: {dataset_name} ===")
        
        # Métadonnées par défaut
        default_metadata = {
            "export_timestamp": datetime.now().isoformat(),
            "dataset_name": dataset_name,
            "shape": df.shape,
            "columns": list(df.columns),
            "data_types": df.dtypes.to_dict(),
            "memory_usage_mb": df.memory_usage(deep=True).sum() / 1024 / 1024,
            "null_counts": df.isnull().sum().to_dict(),
            "unique_counts": df.nunique().to_dict()
        }
        
        # Fusion avec les métadonnées fournies
        if metadata:
            default_metadata.update(metadata)
        
        # Création du répertoire de sortie
        output_directory = output_dir or self.export_config["output_directory"]
        Path(output_directory).mkdir(parents=True, exist_ok=True)
        
        # Export des métadonnées
        timestamp = datetime.now().strftime(self.export_config["timestamp_format"])
        metadata_filename = f"{self.export_config['filename_prefix']}_{dataset_name}_metadata_{timestamp}.json"
        metadata_path = os.path.join(output_directory, metadata_filename)
        
        try:
            with open(metadata_path, 'w', encoding=self.export_config["encoding"]) as f:
                json.dump(default_metadata, f, indent=2, ensure_ascii=False, default=str)
            logger.info(f"📋 Métadonnées exportées: {metadata_path}")
        except Exception as e:
            logger.error(f"❌ Erreur export métadonnées: {e}")
        
        # Export du dataset principal
        exported_files = self.export_dataset(df, dataset_name, output_dir=output_directory)
        
        # Ajout du chemin des métadonnées
        exported_files["metadata"] = metadata_path
        
        return exported_files
    
    def export_chunked(self, df: pd.DataFrame, dataset_name: str, 
                       chunk_size: int = None, output_dir: str = None) -> Dict[str, List[str]]:
        """
        Export par chunks pour les gros datasets
        
        Args:
            df: DataFrame à exporter
            dataset_name: Nom du dataset
            chunk_size: Taille des chunks
            output_dir: Répertoire de sortie
            
        Returns:
            Dict avec les chemins des fichiers exportés par chunk
        """
        logger.info(f"📦 === EXPORT PAR CHUNKS: {dataset_name} ===")
        
        chunk_size = chunk_size or self.export_config["chunk_size"]
        output_directory = output_dir or self.export_config["output_directory"]
        
        # Création du répertoire de sortie
        Path(output_directory).mkdir(parents=True, exist_ok=True)
        
        # Calcul du nombre de chunks
        total_rows = len(df)
        num_chunks = (total_rows + chunk_size - 1) // chunk_size
        
        logger.info(f"📊 Export de {total_rows} lignes en {num_chunks} chunks de {chunk_size}")
        
        exported_chunks = {}
        timestamp = datetime.now().strftime(self.export_config["timestamp_format"])
        
        for chunk_num in range(num_chunks):
            start_idx = chunk_num * chunk_size
            end_idx = min((chunk_num + 1) * chunk_size, total_rows)
            
            chunk_df = df.iloc[start_idx:end_idx]
            chunk_name = f"{dataset_name}_chunk_{chunk_num + 1:03d}"
            
            logger.info(f"📦 Export chunk {chunk_num + 1}/{num_chunks} ({start_idx+1}-{end_idx})")
            
            # Export du chunk
            chunk_files = self.export_dataset(
                chunk_df, 
                chunk_name, 
                output_dir=output_directory
            )
            
            exported_chunks[chunk_name] = chunk_files
        
        # Création d'un fichier d'index des chunks
        index_data = {
            "dataset_name": dataset_name,
            "total_rows": total_rows,
            "chunk_size": chunk_size,
            "num_chunks": num_chunks,
            "export_timestamp": timestamp,
            "chunks": exported_chunks
        }
        
        index_filename = f"{self.export_config['filename_prefix']}_{dataset_name}_chunks_index_{timestamp}.json"
        index_path = os.path.join(output_directory, index_filename)
        
        try:
            with open(index_path, 'w', encoding=self.export_config["encoding"]) as f:
                json.dump(index_data, f, indent=2, ensure_ascii=False, default=str)
            logger.info(f"📋 Index des chunks exporté: {index_path}")
        except Exception as e:
            logger.error(f"❌ Erreur export index chunks: {e}")
        
        exported_chunks["index"] = index_path
        
        return exported_chunks
    
    def generate_export_report(self, dataset_name: str = None, output_path: str = None) -> str:
        """
        Génère un rapport d'export complet
        
        Args:
            dataset_name: Nom du dataset (si None, utilise le dernier exporté)
            output_path: Chemin de sauvegarde (optionnel)
            
        Returns:
            Contenu du rapport
        """
        if dataset_name is None:
            if not self.export_history:
                return "❌ Aucun historique d'export disponible"
            dataset_name = list(self.export_stats.keys())[-1]
        
        if dataset_name not in self.export_stats:
            return f"❌ Dataset '{dataset_name}' non trouvé"
        
        stats = self.export_stats[dataset_name]
        
        # Recherche de l'historique correspondant
        export_record = None
        for record in self.export_history:
            if record["dataset_name"] == dataset_name:
                export_record = record
                break
        
        if not export_record:
            return f"❌ Historique pour '{dataset_name}' non trouvé"
        
        # Génération du rapport
        report_content = []
        report_content.append("# " + "="*80)
        report_content.append("# RAPPORT D'EXPORT MULTI-FORMATS")
        report_content.append("# " + "="*80)
        report_content.append(f"# Dataset: {dataset_name}")
        report_content.append(f"# Date: {export_record['timestamp']}")
        report_content.append(f"# Durée: {export_record['duration_seconds']:.2f}s")
        report_content.append("# " + "="*80 + "\n")
        
        # Résumé exécutif
        report_content.append("## RÉSUMÉ EXÉCUTIF")
        report_content.append(f"**Formats demandés:** {', '.join(export_record['formats'])}")
        report_content.append(f"**Export réussis:** {stats['successful_exports']}/{stats['total_formats']}")
        report_content.append(f"**Taux de succès:** {stats['success_rate']:.1%}")
        report_content.append(f"**Forme du dataset:** {export_record['dataset_shape'][0]} lignes × {export_record['dataset_shape'][1]} colonnes")
        report_content.append(f"**Utilisation mémoire:** {export_record['memory_usage_mb']:.2f} MB")
        report_content.append("")
        
        # Détails par format
        report_content.append("## DÉTAILS PAR FORMAT")
        for format_type, file_path in export_record['exported_files'].items():
            if str(file_path).startswith("ERROR"):
                report_content.append(f"### ❌ {format_type.upper()}")
                report_content.append(f"**Statut:** Échec")
                report_content.append(f"**Erreur:** {file_path}")
            else:
                report_content.append(f"### ✅ {format_type.upper()}")
                report_content.append(f"**Statut:** Succès")
                report_content.append(f"**Fichier:** {file_path}")
                
                # Informations sur le fichier
                if os.path.exists(file_path):
                    file_size = os.path.getsize(file_path) / 1024 / 1024  # MB
                    report_content.append(f"**Taille:** {file_size:.2f} MB")
            
            report_content.append("")
        
        # Statistiques de performance
        report_content.append("## PERFORMANCE")
        report_content.append(f"**Temps total d'export:** {export_record['duration_seconds']:.2f}s")
        report_content.append(f"**Temps moyen par format:** {export_record['duration_seconds'] / len(export_record['formats']):.2f}s")
        report_content.append("")
        
        # Recommandations
        report_content.append("## RECOMMANDATIONS")
        if stats['success_rate'] >= 0.9:
            report_content.append("✅ **Excellent** - Tous les formats ont été exportés avec succès")
        elif stats['success_rate'] >= 0.7:
            report_content.append("✅ **Bon** - La plupart des formats ont été exportés")
        elif stats['success_rate'] >= 0.5:
            report_content.append("⚠️ **Moyen** - Certains formats ont échoué")
        else:
            report_content.append("❌ **Faible** - Beaucoup de formats ont échoué")
        
        report_content.append("")
        report_content.append("# " + "="*80)
        report_content.append("# FIN DU RAPPORT")
        report_content.append("# " + "="*80)
        
        report_text = "\n".join(report_content)
        
        # Sauvegarde si un chemin est fourni
        if output_path:
            try:
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(report_text)
                logger.info(f"📄 Rapport d'export sauvegardé: {output_path}")
            except Exception as e:
                logger.error(f"❌ Erreur sauvegarde rapport: {e}")
        
        return report_text
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Retourne les statistiques d'export globales"""
        if not self.export_stats:
            return {"message": "Aucune statistique d'export disponible"}
        
        total_datasets = len(self.export_stats)
        total_exports = sum(stats["total_formats"] for stats in self.export_stats.values())
        total_successful = sum(stats["successful_exports"] for stats in self.export_stats.values())
        
        return {
            "total_datasets_exported": total_datasets,
            "total_formats_attempted": total_exports,
            "total_successful_exports": total_successful,
            "global_success_rate": total_successful / total_exports if total_exports > 0 else 0,
            "datasets": self.export_stats,
            "export_history_count": len(self.export_history)
        }
